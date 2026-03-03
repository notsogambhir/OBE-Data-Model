"""
CO Attainment Calculator (Pandas + OpenPyXL version)
=====================================================
Calculates Course Outcome (CO) attainment levels from student marks.
Uses pandas for data manipulation and openpyxl for styled Excel output.

Requirements:
    pip install pandas openpyxl numpy

Usage:
    python co_attainment_pandas.py "Input 1.xlsx" "Input 2.xlsx" "Input 3.xlsx"
"""

import sys
import os
import re
import numpy as np
import pandas as pd
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────
# SECTION 1: Parse OBE Details (Thresholds, Weights, Levels)
# ─────────────────────────────────────────────────────────────

def parse_obe_details(xls):
    """
    Read the first sheet (OBE Details) and extract configuration.

    Returns:
        target_pct (float): Course target percentage (e.g. 60.0)
        weights (OrderedDict): Category → weight (e.g. {'Internal': 0.4, 'External': 0.6})
        levels (list): [(level_num, threshold), ...] sorted descending
    """
    df = pd.read_excel(xls, sheet_name=0, header=None)

    target_pct = 60.0
    weights = OrderedDict()
    levels = []
    mode = None

    for _, row in df.iterrows():
        c0 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        c1 = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ''

        if c0.lower() == 'threshold' and c1:
            try:
                target_pct = float(c1)
            except ValueError:
                pass
            continue

        if c0.lower() == 'types' and c1.lower() == 'weightages':
            mode = 'weights'
            continue
        if 'co score' in c0.lower() and 'student' in c1.lower():
            mode = 'levels'
            continue

        if mode == 'weights' and c0 and c1:
            try:
                w = float(c1)
            except ValueError:
                continue
            c0_lower = c0.lower()
            if 'internal' in c0_lower or 'avg' in c0_lower:
                weights['Internal'] = w
            elif 'external' in c0_lower or 'ete' in c0_lower:
                weights['External'] = w
            elif 'assign' in c0_lower or 'asn' in c0_lower:
                weights['Assignment'] = w
            else:
                weights[c0] = w

        if mode == 'levels' and c0 and c1:
            try:
                levels.append((int(float(c0)), float(c1)))
            except ValueError:
                pass

    levels.sort(key=lambda x: x[0], reverse=True)
    if not weights:
        weights = OrderedDict([('Internal', 0.4), ('External', 0.6)])
    if not levels:
        levels = [(3, 0.8), (2, 0.7), (1, 0.6)]

    return target_pct, weights, levels


# ─────────────────────────────────────────────────────────────
# SECTION 2: Discover Exams (Mapping + Result pairs)
# ─────────────────────────────────────────────────────────────

def discover_exams(sheet_names):
    """
    Scan sheet names for Mapping+Result pairs and categorise them.

    Returns:
        list of dicts: {name, map_sheet, res_sheet, category}
    """
    exams = []
    seen = set()

    for s in sheet_names:
        if 'Mapping' not in s:
            continue
        m = s.replace(' Ques Mapping', '').replace(' Mapping', '').strip()
        if m in seen:
            continue
        res = next((s2 for s2 in sheet_names if s2.startswith(m) and 'Result' in s2), None)
        if not res:
            continue

        m_upper = m.upper()
        if m_upper.startswith('ETE') or 'EXTERNAL' in m_upper:
            cat = 'External'
        elif m_upper.startswith('ASN') or 'ASSIGN' in m_upper:
            cat = 'Assignment'
        else:
            cat = 'Internal'

        exams.append({'name': m, 'map_sheet': s, 'res_sheet': res, 'category': cat})
        seen.add(m)

    return exams


# ─────────────────────────────────────────────────────────────
# SECTION 3: Parse Mapping & Result Sheets with Pandas
# ─────────────────────────────────────────────────────────────

def parse_mapping(xls, sheet_name):
    """
    Parse a question mapping sheet into a clean DataFrame.

    Returns:
        map_df (DataFrame): Columns = [Q_Id, Max Marks, CO1, CO2, ...]
        co_cols (list): Sorted list of CO column names found
    """
    # Read WITHOUT dtype=str so pandas can auto-detect types
    df = pd.read_excel(xls, sheet_name=sheet_name)
    df.columns = [str(c).strip() for c in df.columns]

    # Identify CO columns (match CO followed by digits)
    co_cols = sorted([c for c in df.columns if re.match(r'^CO\d+$', c, re.IGNORECASE)],
                     key=lambda x: int(re.search(r'\d+', x).group()))
    # Standardise CO column names to uppercase
    rename_map = {c: c.upper() for c in co_cols}
    df.rename(columns=rename_map, inplace=True)
    co_cols = [c.upper() for c in co_cols]

    # Identify Q_Id and Max Marks columns
    q_col = df.columns[0]  # First column is always question ID
    m_col = df.columns[1]  # Second column is always max marks

    # Clean up
    df[q_col] = df[q_col].astype(str).str.strip()
    df[m_col] = pd.to_numeric(df[m_col], errors='coerce').fillna(0)

    # CO columns may be bool (True/False) or numeric (1/0) depending on Excel format
    for co in co_cols:
        # Convert booleans first, then coerce anything else to numeric
        df[co] = df[co].apply(lambda v: int(v) if isinstance(v, bool) else v)
        df[co] = pd.to_numeric(df[co], errors='coerce').fillna(0).astype(int)

    # Filter: keep only questions with max_marks > 0 and at least one CO mapped
    df = df[df[m_col] > 0].copy()
    df = df[df[co_cols].sum(axis=1) > 0].copy()

    # Rename for consistency
    df = df.rename(columns={q_col: 'Q_Id', m_col: 'Max_Marks'})

    return df[['Q_Id', 'Max_Marks'] + co_cols].reset_index(drop=True), co_cols


def parse_results(xls, sheet_name):
    """
    Parse a student result sheet into a clean DataFrame.

    Returns:
        DataFrame with columns: [Roll, Name, <question_cols>...]
        All marks are float or NaN (NaN = unattempted)
    """
    df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    # Find roll number and name columns
    roll_col = next((c for c in df.columns if 'admission' in c.lower() or 'roll' in c.lower()), None)
    name_col = next((c for c in df.columns if 'name' in c.lower() and 'student' in c.lower()), None)

    if roll_col is None:
        return pd.DataFrame()

    # Identify question columns (exclude metadata)
    meta_keywords = {'sr.no', 'sr no', 'admission', 'roll', 'name', 'course',
                     'exam', 'total', 'max', 'maximum'}
    q_cols = [c for c in df.columns
              if c.strip() and not any(kw in c.lower() for kw in meta_keywords)]

    # Build clean DataFrame
    result = pd.DataFrame()
    result['Roll'] = df[roll_col].astype(str).str.strip().str.replace('\xa0', '', regex=False)
    result['Name'] = df[name_col].astype(str).str.strip().str.replace('\xa0', '', regex=False) if name_col else ''

    # Convert marks: 'U', '', 'AB' → NaN; numeric strings → float
    for qc in q_cols:
        vals = df[qc].astype(str).str.strip().str.upper()
        vals = vals.replace({'U': np.nan, '': np.nan, 'AB': np.nan, 'NAN': np.nan})
        result[qc] = pd.to_numeric(vals, errors='coerce')

    result = result[result['Roll'].str.len() > 0].reset_index(drop=True)
    return result


# ─────────────────────────────────────────────────────────────
# SECTION 4: Core CO Calculation Engine (Vectorised with Pandas)
# ─────────────────────────────────────────────────────────────

def calc_exam_co_pct(result_df, map_df, co_name):
    """
    Calculate CO percentage for ALL students in one exam (vectorised).

    For each student:
        CO% = sum(obtained on attempted mapped questions) /
              sum(max marks on attempted mapped questions) × 100

    Parameters:
        result_df (DataFrame): Student results with question columns
        map_df (DataFrame): Question mapping with Q_Id, Max_Marks, CO columns
        co_name (str): Which CO to calculate (e.g. 'CO1')

    Returns:
        Series: CO percentage per student (indexed by row), NaN where no data
    """
    if co_name not in map_df.columns:
        return pd.Series(np.nan, index=result_df.index)

    # Get questions mapped to this CO
    mapped = map_df[map_df[co_name] > 0]
    if mapped.empty:
        return pd.Series(np.nan, index=result_df.index)

    obtained_total = pd.Series(0.0, index=result_df.index)
    max_total = pd.Series(0.0, index=result_df.index)

    for _, q_row in mapped.iterrows():
        q_id = q_row['Q_Id']
        max_marks = q_row['Max_Marks']

        if q_id not in result_df.columns:
            continue

        marks = result_df[q_id]
        attempted = marks.notna()  # Boolean mask: True where student attempted

        obtained_total += marks.fillna(0) * attempted
        max_total += max_marks * attempted

    # Calculate percentage; NaN where max_total == 0 (no attempted questions)
    pct = np.where(max_total > 0, (obtained_total / max_total) * 100, np.nan)
    return pd.Series(pct, index=result_df.index)


# ─────────────────────────────────────────────────────────────
# SECTION 5: Main Processing Pipeline
# ─────────────────────────────────────────────────────────────

def process_file(filepath):
    """
    End-to-end processing of one input Excel file.

    Returns:
        student_master (DataFrame): Per-student CO percentages and target check
        attainment_df (DataFrame): Class-level CO attainment summary
        meta (dict): Configuration used
    """
    print(f"\n{'='*60}")
    print(f"  Processing: {os.path.basename(filepath)}")
    print(f"{'='*60}")

    xls = pd.ExcelFile(filepath)
    sheet_names = xls.sheet_names

    # ── Step 1: Parse OBE Config ──
    target_pct, weights, levels = parse_obe_details(xls)
    print(f"  Target: {target_pct}%")
    print(f"  Weights: {dict(weights)}")
    print(f"  Levels: {levels}")

    # ── Step 2: Discover exams ──
    exams = discover_exams(sheet_names)
    print(f"  Exams: {[e['name'] + ' (' + e['category'] + ')' for e in exams]}")

    # ── Step 3: Parse data for each exam ──
    exam_data = []
    all_cos = set()
    all_students = pd.DataFrame()

    for exam in exams:
        map_df, cos = parse_mapping(xls, exam['map_sheet'])
        res_df = parse_results(xls, exam['res_sheet'])

        if map_df.empty:
            print(f"    Skipping {exam['name']}: no valid question mappings")
            continue

        # Collect student roster
        if not res_df.empty:
            roster = res_df[['Roll', 'Name']].drop_duplicates(subset='Roll')
            all_students = pd.concat([all_students, roster]).drop_duplicates(subset='Roll')

        all_cos.update(cos)
        exam_data.append({
            'name': exam['name'],
            'category': exam['category'],
            'map_df': map_df,
            'res_df': res_df,
            'cos': cos
        })

    all_cos = sorted(all_cos)
    all_students = all_students.reset_index(drop=True)
    print(f"  COs: {all_cos}")
    print(f"  Students: {len(all_students)}")

    # ── Step 4: Per-student, per-exam CO percentages ──
    # Build DataFrames: one per exam, indexed by Roll
    exam_co_pcts = {}  # {exam_name: DataFrame with CO columns}
    for ed in exam_data:
        res = ed['res_df'].copy()
        if res.empty:
            continue
        res = res.set_index('Roll')
        pct_df = pd.DataFrame(index=res.index)
        for co in all_cos:
            pct_df[co] = calc_exam_co_pct(res, ed['map_df'], co)
        exam_co_pcts[ed['name']] = pct_df

    # ── Step 5: Weighted final CO% per student ──
    cat_exams = {}
    for ed in exam_data:
        cat_exams.setdefault(ed['category'], []).append(ed['name'])

    student_master = all_students.set_index('Roll').copy()

    for co in all_cos:
        weighted = pd.Series(0.0, index=student_master.index)

        for cat, w in weights.items():
            cat_exam_names = cat_exams.get(cat, [])
            if not cat_exam_names:
                continue

            # Collect CO% from all exams in this category
            cat_pcts = pd.DataFrame(index=student_master.index)
            for ename in cat_exam_names:
                if ename in exam_co_pcts:
                    series = exam_co_pcts[ename].reindex(student_master.index)[co]
                    cat_pcts[ename] = series

            # Average across exams, then multiply by weight
            if not cat_pcts.empty:
                cat_avg = cat_pcts.mean(axis=1, skipna=True)
                weighted += cat_avg.fillna(0) * w

        # Store final CO%
        student_master[f'{co} %'] = weighted.replace(0, np.nan).round(2)
        student_master[f'{co} Target'] = np.where(
            student_master[f'{co} %'].notna(),
            np.where(student_master[f'{co} %'] >= target_pct, 'Yes', 'No'),
            'N/A'
        )

    # ── Step 6: Class-level aggregation ──
    summary_rows = []
    for co in all_cos:
        col = f'{co} %'
        valid = student_master[col].dropna()
        total = len(valid)
        passed = (valid >= target_pct).sum()
        rate = passed / total if total > 0 else 0.0

        level = 0
        for lv, th in levels:
            if rate >= th:
                level = lv
                break

        summary_rows.append({
            'Course Outcome': co,
            'Students Attempted': total,
            'Students Meeting Target': passed,
            'Success Rate %': round(rate * 100, 2),
            'Attainment Level': level
        })

    attainment_df = pd.DataFrame(summary_rows)

    # ── Print summary ──
    print(f"\n  --- CO Attainment Summary ---")
    print(f"  {'CO':<8} {'Attempted':>10} {'Met Target':>12} {'Success%':>10} {'Level':>6}")
    print(f"  {'-'*50}")
    for _, r in attainment_df.iterrows():
        print(f"  {r['Course Outcome']:<8} {r['Students Attempted']:>10} "
              f"{r['Students Meeting Target']:>12} {r['Success Rate %']:>9.2f}% "
              f"{r['Attainment Level']:>6}")

    meta = {
        'target': target_pct,
        'weights': dict(weights),
        'levels': levels,
        'cos': all_cos
    }

    return student_master.reset_index(), attainment_df, meta


# ─────────────────────────────────────────────────────────────
# SECTION 6: Styled Excel Output (openpyxl)
# ─────────────────────────────────────────────────────────────

def write_styled_output(filepath, student_df, attainment_df, meta):
    """
    Write results to a professionally styled .xlsx file using openpyxl.
    """
    wb = Workbook()

    # ── Styles ──
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    level_fills = {
        3: PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
        2: PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),
        1: PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),
        0: PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
    }
    level_fonts = {
        3: Font(bold=True, color='FFFFFF'),
        2: Font(bold=True, color='000000'),
        1: Font(bold=True, color='000000'),
        0: Font(bold=True, color='FFFFFF'),
    }

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

    # ════════ Sheet 1: Student Details ════════
    ws1 = wb.active
    ws1.title = 'Student Details'

    headers = list(student_df.columns)
    ws1.append(headers)
    style_header(ws1)

    for _, row_data in student_df.iterrows():
        ws1.append(list(row_data.values))

    # Apply conditional formatting to Target columns
    target_cols = [i + 1 for i, h in enumerate(headers) if 'Target' in str(h)]
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        for cell in row:
            cell.alignment = data_align
            cell.border = thin_border
            if cell.column in target_cols:
                if cell.value == 'Yes':
                    cell.fill = green_fill
                elif cell.value == 'No':
                    cell.fill = red_fill

    auto_width(ws1)
    ws1.freeze_panes = 'C2'  # Freeze Roll + Name columns

    # ════════ Sheet 2: Attainment Summary ════════
    ws2 = wb.create_sheet('Attainment Summary')

    headers2 = list(attainment_df.columns)
    ws2.append(headers2)
    style_header(ws2)

    for _, row_data in attainment_df.iterrows():
        ws2.append(list(row_data.values))

    # Style level cells with colour coding
    level_col = headers2.index('Attainment Level') + 1
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.alignment = data_align
            cell.border = thin_border
            if cell.column == level_col:
                lv = int(cell.value) if cell.value is not None else 0
                cell.fill = level_fills.get(lv, PatternFill())
                cell.font = level_fonts.get(lv, Font())

    auto_width(ws2)

    # ════════ Sheet 3: Configuration ════════
    ws3 = wb.create_sheet('Configuration')
    ws3.append(['Parameter', 'Value'])
    style_header(ws3)

    config_rows = [('Target Percentage', meta['target'])]
    for cat, w in meta['weights'].items():
        config_rows.append((f'Weight: {cat}', w))
    for lv, th in meta['levels']:
        config_rows.append((f'Level {lv} Threshold', th))

    for param, val in config_rows:
        ws3.append([param, val])

    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in row:
            cell.alignment = data_align
            cell.border = thin_border

    auto_width(ws3)

    # Save
    wb.save(filepath)
    print(f"\n  Output saved to: {filepath}")


# ─────────────────────────────────────────────────────────────
# SECTION 7: Main Entry Point
# ─────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        files = sorted([f for f in os.listdir('.')
                        if f.lower().startswith('input') and f.lower().endswith('.xlsx')])
        if not files:
            print("Usage: python co_attainment_pandas.py <file1.xlsx> [file2.xlsx] ...")
            sys.exit(1)
    else:
        files = sys.argv[1:]

    for filepath in files:
        if not os.path.exists(filepath):
            print(f"ERROR: File not found: {filepath}")
            continue
        try:
            student_df, attainment_df, meta = process_file(filepath)
            base = os.path.splitext(os.path.basename(filepath))[0]
            out_path = os.path.join(os.path.dirname(filepath) or '.', f"Output_{base}.xlsx")
            write_styled_output(out_path, student_df, attainment_df, meta)
        except Exception as e:
            print(f"ERROR processing {filepath}: {e}")
            import traceback
            traceback.print_exc()

    print(f"\n{'='*60}")
    print("  All files processed!")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
